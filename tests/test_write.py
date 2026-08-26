from decimal import Decimal

import openpyxl
import pytest

from po2xlsx.cli import main
from po2xlsx.write import (
    exact_decimal_serialisation,
    number_format,
    write_workbook,
)

from conftest import SAMPLE, TEMPLATE, TEMPLATE_COLUMNS


def read(path):
    return openpyxl.load_workbook(path).active


def column(sheet, heading):
    return list(sheet[1]).index(
        next(c for c in sheet[1] if c.value == heading)
    ) + 1


def test_column_order_comes_from_the_template(sample, template, tmp_path):
    out = tmp_path / "out.xlsx"
    write_workbook([sample], template, out)
    sheet = read(out)
    assert [c.value for c in sheet[1]] == TEMPLATE_COLUMNS
    assert sheet.max_row == 1 + len(sample.items)


def test_template_drives_order_even_when_reordered(sample, tmp_path):
    """Reordering the template must reorder the output with no code change."""
    reordered = TEMPLATE_COLUMNS[::-1]
    workbook = openpyxl.Workbook()
    workbook.active.append(reordered)
    path = tmp_path / "reordered.xlsx"
    workbook.save(path)

    out = tmp_path / "out.xlsx"
    write_workbook([sample], path, out)
    sheet = read(out)
    assert [c.value for c in sheet[1]] == reordered
    assert sheet.cell(row=2, column=column(sheet, "UPC")).value == "0091120032129"


def test_unknown_template_column_is_reported_and_left_blank(sample, tmp_path):
    workbook = openpyxl.Workbook()
    workbook.active.append(["FILENAME", "SKU", "SOMETHING NEW"])
    path = tmp_path / "extra.xlsx"
    workbook.save(path)

    out = tmp_path / "out.xlsx"
    warnings = write_workbook([sample], path, out)
    assert any("SOMETHING NEW" in w for w in warnings)
    assert read(out).cell(row=2, column=3).value is None


# --- type preservation ----------------------------------------------------


def test_leading_zeros_survive_a_round_trip(sample, template, tmp_path):
    out = tmp_path / "out.xlsx"
    write_workbook([sample], template, out)
    sheet = read(out)

    for heading, expected in [
        ("UPC", "0091120032129"),
        ("Vendor Part #", "04212"),
        ("SKU", "71097188"),
        ("Dept", "210"),
    ]:
        cell = sheet.cell(row=2, column=column(sheet, heading))
        assert cell.value == expected
        assert isinstance(cell.value, str)
        assert cell.number_format == "@"       # not merely a str: Excel-safe


def test_money_keeps_its_printed_scale(sample, template, tmp_path):
    """82913.600 must reach the file with its trailing zeros, and must not be
    written through a float."""
    out = tmp_path / "out.xlsx"
    write_workbook([sample], template, out)

    sheet = read(out)
    cell = sheet.cell(row=3, column=column(sheet, "EXT Cost"))
    assert cell.value == pytest.approx(82913.6)
    assert cell.number_format == "0.000"

    import zipfile

    with zipfile.ZipFile(out) as archive:
        xml = archive.read("xl/worksheets/sheet1.xml").decode()
    assert "<v>82913.600</v>" in xml
    assert "60000000001" not in xml            # openpyxl's default float leak


@pytest.mark.parametrize(
    "value, expected",
    [
        (Decimal("82913.600"), "0.000"),
        (Decimal("149.99"), "0.00"),
        (Decimal("0.0991"), "0.0000"),
        (Decimal("68"), "0"),
    ],
)
def test_number_format_follows_source_scale(value, expected):
    assert number_format(value) == expected


def test_blank_stays_blank(mini, template, tmp_path):
    out = tmp_path / "out.xlsx"
    write_workbook([mini], template, out)
    sheet = read(out)
    assert sheet.cell(row=3, column=column(sheet, "Dept")).value is None
    assert sheet.cell(row=4, column=column(sheet, "Vendor Part #")).value is None
    assert sheet.cell(row=2, column=column(sheet, "REF MASTER PO#")).value is None


# --- Total Cost -----------------------------------------------------------


def test_total_cost_invoice_mode_repeats_the_document_total(sample, template, tmp_path):
    out = tmp_path / "out.xlsx"
    write_workbook([sample], template, out, total_cost_mode="invoice")
    sheet = read(out)
    index = column(sheet, "Total Cost")
    values = [sheet.cell(row=r, column=index).value for r in range(2, sheet.max_row + 1)]
    assert values == [pytest.approx(447455.12)] * len(sample.items)


def test_total_cost_ext_mode_duplicates_the_row(sample, template, tmp_path):
    out = tmp_path / "out.xlsx"
    write_workbook([sample], template, out, total_cost_mode="ext")
    sheet = read(out)
    total = column(sheet, "Total Cost")
    ext = column(sheet, "EXT Cost")
    for row in range(2, sheet.max_row + 1):
        assert sheet.cell(row=row, column=total).value == (
            sheet.cell(row=row, column=ext).value
        )


def test_unknown_total_cost_mode_is_rejected(sample, template, tmp_path):
    with pytest.raises(ValueError):
        write_workbook([sample], template, tmp_path / "out.xlsx", total_cost_mode="x")


# --- batch ----------------------------------------------------------------


def test_batch_writes_one_workbook_with_filename_per_row(
    sample, mini, template, tmp_path
):
    out = tmp_path / "combined.xlsx"
    write_workbook([sample, mini], template, out)
    sheet = read(out)
    assert sheet.max_row == 1 + len(sample.items) + len(mini.items)
    names = [
        sheet.cell(row=r, column=column(sheet, "FILENAME")).value
        for r in range(2, sheet.max_row + 1)
    ]
    assert names.count("purchase_order_sample.txt") == len(sample.items)
    assert names.count("mini_po.txt") == len(mini.items)


def test_cli_reports_validation_errors_and_still_writes(
    template, tmp_path, fixtures_dir, capsys
):
    out = tmp_path / "out.xlsx"
    code = main([
        str(fixtures_dir / "broken_po.txt"),
        "--template", str(template), "--out", str(out), "--quiet",
    ])
    assert code == 1                            # errors reported, not swallowed
    assert out.exists()                         # output still produced
    assert "EXT QTY 12 != CTNS 5 x CSPK 2" in capsys.readouterr().err


@pytest.mark.skipif(not TEMPLATE.exists(), reason="provided template not present")
def test_against_the_provided_template(sample, tmp_path):
    out = tmp_path / "out.xlsx"
    assert write_workbook([sample], TEMPLATE, out) == []
    assert [c.value for c in read(out)[1]] == TEMPLATE_COLUMNS


# --- regressions from PR review -------------------------------------------


def test_non_utf8_input_is_decoded_but_reported(template, tmp_path, fixtures_dir, capsys):
    """cp1252 bytes must not be silently replaced with U+FFFD -- a corrupted
    UPC that still reports success is the one failure this must not have."""
    source = (fixtures_dir / "mini_po.txt").read_bytes().replace(
        b"TALL BOX", b"TALL B\xd8X"
    )
    path = tmp_path / "cp1252.txt"
    path.write_bytes(source)

    out = tmp_path / "out.xlsx"
    code = main([str(path), "--template", str(template), "--out", str(out), "--quiet"])
    assert code == 0
    assert "decoded as cp1252" in capsys.readouterr().err

    sheet = read(out)
    descriptions = [
        sheet.cell(row=r, column=column(sheet, "Description")).value
        for r in range(2, sheet.max_row + 1)
    ]
    assert "TALL BØX" in descriptions           # decoded, not replaced
    assert not any("�" in d for d in descriptions)


def test_undecodable_input_fails_loudly(template, tmp_path, capsys):
    path = tmp_path / "binary.txt"
    path.write_bytes(b"\x81\x8d\x8f\x90\x9d" * 4)

    code = main([str(path), "--template", str(template), "--out",
                 str(tmp_path / "out.xlsx"), "--quiet"])
    assert code == 2
    assert "could not decode" in capsys.readouterr().err


def test_summary_names_what_totals_cross_checked(template, tmp_path, capsys):
    """A document printing one total reconciles as cleanly as one printing
    four; the summary has to say which columns were actually tied out."""
    main([str(SAMPLE), "--template", str(template),
          "--out", str(tmp_path / "out.xlsx")])
    out = capsys.readouterr().out
    assert "TOTALS cross-checked ext_cost, ctns, ext_qty, kilograms" in out


def test_precision_patch_reports_when_it_cannot_take_effect(monkeypatch):
    """Exact Decimal output rides on an openpyxl internal, so an upgrade can
    take it away."""
    from openpyxl.cell import _writer

    monkeypatch.delattr(_writer, "safe_string")
    with exact_decimal_serialisation() as warnings:
        pass
    assert any("may not keep their printed scale" in w for w in warnings)


def test_precision_patch_is_removed_again_afterwards():
    """It edits a module global: leaving it installed would leak into any other
    openpyxl use in the same process."""
    from openpyxl.cell import _writer

    before = _writer.safe_string
    with exact_decimal_serialisation() as warnings:
        assert _writer.safe_string is not before
        assert warnings == []
    assert _writer.safe_string is before
