"""Records -> workbook, driven by the template's own header row.

We read row 1, match each heading to a producer, and fill beneath it.
"""

from __future__ import annotations

import re
from contextlib import contextmanager
from decimal import Decimal
from pathlib import Path

import openpyxl
from openpyxl.cell import _writer
from openpyxl.styles import Font, PatternFill

from .parse import PurchaseOrder
from .validate import ERROR, Issue, match_totals

__all__ = [
    "TOTAL_COST_MODES",
    "exact_decimal_serialisation",
    "number_format",
    "text_format",
    "write_workbook",
]

#: `Total Cost` is ambiguous in the brief. We support both readings, and the
#: choice is recorded in the README.
TOTAL_COST_MODES = ("invoice", "ext")

#: Sheet appended alongside the line items when validation findings are passed
#: in. Sheet1's schema is untouched, so the output contract still holds.
VALIDATION_SHEET = "Validation"
VALIDATION_COLUMNS = ("Severity", "File", "Line", "Column", "Message")

#: Excel's own bad-cell and neutral-cell colours, so a flagged cell reads the
#: way a spreadsheet user already expects it to.
ERROR_FILL = PatternFill("solid", fgColor="FFFFC7CE")
WARNING_FILL = PatternFill("solid", fgColor="FFFFEB9C")

#: Excel's "Text" number format. Set explicitly so a value that *is* a
#: str is still protected.
TEXT = "@"


def text_format(value: str | None) -> tuple[str | None, str]:
    return (value or None), TEXT


def number_format(value: Decimal) -> str:
    """A format that preserves the source's printed scale.

    `82913.600` is parsed as Decimal with exponent -3, so it displays as
    `82913.600` rather than `82913.6`.
    """
    exponent = value.as_tuple().exponent
    if not isinstance(exponent, int) or exponent >= 0:
        return "0"
    return "0." + "0" * -exponent


@contextmanager
def exact_decimal_serialisation():
    """Make openpyxl write Decimals verbatim for the duration of a save.

    openpyxl formats every numeric cell with `"%.16g" % value`, which coerces a
    Decimal through float and prints its binary representation error.

    Yields a list of warnings, empty when the patch took effect.
    """
    warnings: list[str] = []
    original = getattr(_writer, "safe_string", None)
    if original is None:
        warnings.append(
            f"openpyxl {openpyxl.__version__} does not expose "
            "cell._writer.safe_string; Decimal values are written through "
            "float and may not keep their printed scale"
        )
        yield warnings
        return

    def safe_string(value):
        # `format(..., "f")` keeps the source's scale: 82913.600 stays 82913.600.
        return format(value, "f") if isinstance(value, Decimal) else original(value)

    _writer.safe_string = safe_string
    try:
        yield warnings
    finally:
        _writer.safe_string = original


def _resolve_total_cost(po: PurchaseOrder) -> Decimal | None:
    """The document-level figure that fills `Total Cost` in `invoice` mode."""
    if po.header.total_invoice_value is not None:
        return po.header.total_invoice_value
    match = match_totals(po)
    # `or` would read a legitimate document total of exactly zero as absent.
    matched = match.matched.get("ext_cost")
    return match.sums.get("ext_cost") if matched is None else matched


#: Normalised template heading -> attribute on Header, or on LineItem.
_HEADER_FIELDS = {
    "BUYER": "buyer",
    "SHIPTERMS": "ship_terms",
    "PO#": "po_number",
    "REFMASTERPO#": "ref_master_po",
    "SHIPDATE": "ship_date",
    "VENDOR": "vendor",
    "SHIPTO": "ship_to",
    "BILLTO": "bill_to",
}
_ITEM_TEXT_FIELDS = {
    "DEPT": "dept",
    "SKU": "sku",
    "UPC": "upc",
    "VENDORPART#": "vendor_part",
    "DESCRIPTION": "description",
}
_ITEM_NUMERIC_FIELDS = {
    "RETAIL": "retail",
    "COST": "cost",
    "EXTCOST": "ext_cost",
    "CTNS": "ctns",
    "CSPK": "cspk",
    "EXTQTY": "ext_qty",
    "CUBE": "cube",
    "KILOGRAMS": "kilograms",
}


def _normalise(heading: object) -> str:
    return re.sub(r"\s+", "", str(heading or "")).upper()


def write_workbook(
    pos: list[PurchaseOrder],
    template_path: str | Path,
    out_path: str | Path,
    total_cost_mode: str = "invoice",
    issues: list[Issue] | None = None,
) -> list[str]:
    """Write every line item of every PO into a copy of the template.

    Pass `issues` to surface validation findings in the workbook.

    Returns warnings about template headings we could not fill.
    """
    if total_cost_mode not in TOTAL_COST_MODES:
        raise ValueError(
            f"total_cost_mode must be one of {TOTAL_COST_MODES}, got {total_cost_mode!r}"
        )

    workbook = openpyxl.load_workbook(template_path)
    sheet = workbook.active
    headings = [_normalise(cell.value) for cell in sheet[1]]

    unknown = [
        cell.value
        for cell, key in zip(sheet[1], headings)
        if key
        and key
        not in {"FILENAME", "TOTALCOST", *_HEADER_FIELDS, *_ITEM_TEXT_FIELDS,
                *_ITEM_NUMERIC_FIELDS}
    ]
    warnings = [
        f"template column {name!r} is not produced by this parser; left blank"
        for name in unknown
    ]

    row = 2
    rows_by_source: dict[tuple[str, int], int] = {}
    for po in pos:
        document_total = _resolve_total_cost(po)
        for item in po.items:
            for index, key in enumerate(headings, start=1):
                value, fmt = _cell(po, item, key, total_cost_mode, document_total)
                cell = sheet.cell(row=row, column=index)
                cell.number_format = fmt
                if value is not None:
                    _set(cell, value, fmt)
            rows_by_source[(po.filename, item.line_no)] = row
            row += 1

    if issues is not None:
        _flag_cells(sheet, headings, rows_by_source, issues)
        headings_text = {
            _normalise(c.value): c.value for c in sheet[1] if c.value
        }
        _append_validation_sheet(workbook, pos, issues, headings_text)

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    with exact_decimal_serialisation() as precision_warnings:
        workbook.save(out_path)
    return warnings + precision_warnings


def _set(cell, value, fmt) -> None:
    """Assign a cell, keeping source text as text.

    openpyxl types any string beginning with "=" as a formula, so a vendor
    description reading `=HYPERLINK(...)` would reach Excel as live code rather
    than as the characters the PO printed.
    """
    cell.value = value
    if fmt == TEXT and isinstance(value, str):
        cell.data_type = "s"


#: Canonical field name -> the normalised template heading that carries it.
_COLUMN_HEADINGS = {
    field: heading
    for mapping in (_ITEM_TEXT_FIELDS, _ITEM_NUMERIC_FIELDS)
    for heading, field in mapping.items()
}


def _flag_cells(sheet, headings, rows_by_source, issues) -> None:
    """Fill the cell each row-level finding blames.
    """
    for issue in issues:
        if issue.line_no is None or issue.column is None:
            continue
        row = rows_by_source.get((issue.filename, issue.line_no))
        heading = _COLUMN_HEADINGS.get(issue.column)
        if row is None or heading is None or heading not in headings:
            continue
        cell = sheet.cell(row=row, column=headings.index(heading) + 1)
        # An error already flagged must not be downgraded by a later warning.
        if issue.severity == ERROR or cell.fill != ERROR_FILL:
            cell.fill = ERROR_FILL if issue.severity == ERROR else WARNING_FILL


def _append_validation_sheet(workbook, pos, issues, headings_text) -> None:
    """Append the findings as their own sheet, one row per finding.

    A file with nothing to report still gets a row: silence and "not checked"
    look identical otherwise.
    """
    sheet = workbook.create_sheet(VALIDATION_SHEET)
    sheet.append(list(VALIDATION_COLUMNS))
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    by_file: dict[str, list[Issue]] = {po.filename: [] for po in pos}
    for issue in issues:
        by_file.setdefault(issue.filename, []).append(issue)

    for filename, found in by_file.items():
        if not found:
            sheet.append(["ok", filename, None, None, "no discrepancies found"])
            continue
        for issue in found:
            key = _COLUMN_HEADINGS.get(issue.column or "")
            sheet.append([
                issue.severity, issue.filename, issue.line_no,
                # The template's own spelling, so the sheet points at a column
                # the reader can actually find in Sheet1.
                headings_text.get(key, issue.column), issue.message,
            ])

    for column, width in zip("ABCDE", (10, 28, 8, 16, 96)):
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"
    workbook.active = 0          # the line items stay the sheet that opens


def _cell(po, item, key, total_cost_mode, document_total):
    """(value, number_format) for one template column. None means leave blank."""
    if key == "FILENAME":
        return text_format(po.filename)
    if key in _HEADER_FIELDS:
        return text_format(getattr(po.header, _HEADER_FIELDS[key]))
    if key in _ITEM_TEXT_FIELDS:
        return text_format(getattr(item, _ITEM_TEXT_FIELDS[key]))
    if key in _ITEM_NUMERIC_FIELDS:
        value = getattr(item, _ITEM_NUMERIC_FIELDS[key])
        return (None, "General") if value is None else (value, number_format(value))
    if key == "TOTALCOST":
        value = item.ext_cost if total_cost_mode == "ext" else document_total
        return (None, "General") if value is None else (value, number_format(value))
    return None, "General"
